import os
import secrets
import csv
import io
import json
import socket
from functools import wraps
from datetime import datetime
from urllib.parse import urlparse

from dotenv import load_dotenv
from flask import Flask, flash, redirect, render_template, request, session, url_for, make_response
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text, func
from sqlalchemy.exc import OperationalError
from sqlalchemy.pool import NullPool
from werkzeug.security import check_password_hash, generate_password_hash

load_dotenv()


def resolve_database_url() -> str:
    database_url = os.getenv("DATABASE_URL", "sqlite:///portal.db")

    # Render provides postgres:// but SQLAlchemy >= 1.4 requires postgresql://
    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql://", 1)

    if database_url.startswith("postgresql://"):
        parsed = urlparse(database_url)
        hostname = parsed.hostname
        if hostname:
            try:
                socket.getaddrinfo(hostname, parsed.port or 5432)
            except socket.gaierror:
                fallback_url = "sqlite:///portal.db"
                print(
                    f"[DB CONFIG WARNING] Unresolvable database host '{hostname}'. "
                    f"Falling back to {fallback_url}."
                )
                return fallback_url

    return database_url

app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", secrets.token_hex(32))
_database_url = resolve_database_url()
app.config["SQLALCHEMY_DATABASE_URI"] = _database_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# Avoid connection-pool exhaustion (SQLAlchemy error e3q8) on small Render DB plans.
if _database_url.startswith("sqlite"):
    app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
        "pool_pre_ping": True,
    }
else:
    engine_opts = {
        "pool_pre_ping": True,
        "pool_recycle": 300,
    }
    if os.getenv("RENDER"):
        # Render free DB plans are sensitive to pooled idle connections.
        # NullPool avoids connection checkout starvation/timeouts (e3q8).
        engine_opts["poolclass"] = NullPool
    app.config["SQLALCHEMY_ENGINE_OPTIONS"] = engine_opts

db = SQLAlchemy(app)


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False, index=True)
    email = db.Column(db.String(120), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(255), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    cbhi_reports = db.relationship("CbhiReport", backref="cbhi_owner", lazy=True, cascade="all, delete-orphan")
    ncd_reports = db.relationship("NcdReport", backref="ncd_owner", lazy=True, cascade="all, delete-orphan")

    def set_password(self, password: str) -> None:
        self.password_hash = generate_password_hash(password)

    def check_password(self, password: str) -> bool:
        return check_password_hash(self.password_hash, password)







COMMUNICABLE_DISEASES = [
    ("1", "Acute diarrhoeal diseases (including gastro enteritis etc.)", "A09"),
    ("2", "Acute poliomyelitis", "A80"),
    ("3", "Acute respiratory infection including influenza and excluding pneumonia", "J00-J06, J10-J11, J20-J22"),
    ("4", "AIDS (as reported by NACO)", "B20-B24"),
    ("5", "Chicken pox", "B01"),
    ("6", "Cholera (lab confirmed)", "A00"),
    ("7", "Corona", "U07.1"),
    ("8", "Diphtheria (lab confirmed)", "A36"),
    ("9", "Acute conjunctivitis / eye infection", "H10"),
    ("10", "Encephalitis", "G04.9"),
    ("11", "Enteric fever (lab. confirmed)", "A01"),
    ("12", "Gonococcal infection", "A54"),
    ("13", "Leptospirosis (lab. confirmed)", "A27"),
    ("14", "Measles (lab. confirmed)", "B05"),
    ("15", "Meningitis (other than bacterial)", "G03"),
    ("16", "Neonatal tetanus (lab. confirmed)", "A33"),
    ("17", "Other STD cases", "A50-A64"),
    ("18", "Rabies", "A82"),
    ("19", "Pneumonia", "J12-J18"),
    ("20", "Pulmonary tuberculosis", "A15"),
    ("21", "Scabies", "B86"),
    ("22", "Scrub typhus (lab. confirmed)", "A75.3"),
    ("23", "Swine flu", "J10"),
    ("24", "Typhus", "A75-A79"),
    ("25", "Tetanus other than neonatal", "A35"),
    ("26", "Viral hepatitis - A (lab. confirmed)", "B15.9"),
    ("27", "Viral hepatitis - B (lab. confirmed)", "B16.9"),
    ("28", "Viral hepatitis - C (lab. confirmed)", "B17.1"),
    ("29", "Viral hepatitis - D (lab. confirmed)", "B17.0"),
    ("30", "Viral hepatitis - E (lab. confirmed)", "B17.2"),
    ("31", "Viral meningitis", "G03.9"),
    ("32", "Whooping cough (lab. confirmed)", "A37"),
]


class CbhiReport(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    health_establishment_name = db.Column(db.String(255), nullable=False)
    postal_address_phone = db.Column(db.Text, nullable=False)
    month = db.Column(db.String(20), nullable=False)
    year = db.Column(db.String(10), nullable=False)
    entries_json = db.Column(db.Text, default="{}")
    approving_authority_name = db.Column(db.String(255), default="")
    approving_authority_designation = db.Column(db.String(255), default="")
    official_email = db.Column(db.String(255), default="")
    official_phone = db.Column(db.String(50), default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)


NON_COMMUNICABLE_DISEASES = [
    ("1", "Accidental injuries", "S00-S99; T00-T14"),
    ("2", "Arsenicosis", "T57.0"),
    ("3", "Asthma", "J45"),
    ("4", "Autoimmune diseases", "M30-M36"),
    ("5", "Bronchitis", "J40"),
    ("6", "Burns", "T20-T32"),
    ("7", "Cancer (oral lip, oral cavity and pharynx)", "C00-C14; D10"),
    ("8", "Cancer (breast)", "C50; D24"),
    ("9", "Cancer (cervix)", "C53; D26"),
    ("10", "Cancer (lung)", "C34; D14.3"),
    ("11", "Cancer (others excluding 7-10)", "C00-D48"),
    ("12", "Cerebrovascular accident", "I60-I69"),
    ("13", "Diabetes mellitus (Type 1)", "E10"),
    ("14", "Diabetes mellitus (Type 2)", "E11"),
    ("15", "Emphysema", "J43"),
    ("16", "Heart diseases (congenital)", "Q20-Q28"),
    ("17", "Heart diseases (ischemic)", "I20-I25"),
    ("18", "Hypertension", "I10-I15"),
    ("19", "Mental disorders", "F10-F19; F99"),
    ("20", "Neurological disorder (chronic)", "G90-G99"),
    ("21", "Neurological disorders (other)", "G00-G89"),
    ("22", "Obesity", "E66.9"),
    ("23", "Other cardiovascular diseases", "I05-I09; I26-I70"),
    ("24", "Others", "-"),
    ("25", "Renal failure (acute)", "N17"),
    ("26", "Renal failure (chronic)", "N18"),
    ("27", "Rheumatic fever", "I00-I02"),
    ("28", "Road traffic accidents", "V01-V89"),
    ("29", "Rubella", "B06"),
    ("30", "Severe mental disorder", "F99"),
    ("31", "Snake bite", "T63.0"),
]


class NcdReport(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    health_establishment_name = db.Column(db.String(255), nullable=False)
    postal_address_phone = db.Column(db.Text, nullable=False)
    month = db.Column(db.String(20), nullable=False)
    year = db.Column(db.String(10), nullable=False)
    entries_json = db.Column(db.Text, default="{}")
    approving_authority_name = db.Column(db.String(255), default="")
    approving_authority_designation = db.Column(db.String(255), default="")
    official_email = db.Column(db.String(255), default="")
    official_phone = db.Column(db.String(50), default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)


def build_cbhi_payload(form_source):
    payload = {}
    for sr_no, _, _ in COMMUNICABLE_DISEASES:
        payload[sr_no] = {
            "general_opd_m": form_source.get(f"cbhi_{sr_no}_general_opd_m", 0, type=int),
            "general_opd_f": form_source.get(f"cbhi_{sr_no}_general_opd_f", 0, type=int),
            "general_opd_tr": form_source.get(f"cbhi_{sr_no}_general_opd_tr", 0, type=int),
            "emergency_opd_m": form_source.get(f"cbhi_{sr_no}_emergency_opd_m", 0, type=int),
            "emergency_opd_f": form_source.get(f"cbhi_{sr_no}_emergency_opd_f", 0, type=int),
            "emergency_opd_tr": form_source.get(f"cbhi_{sr_no}_emergency_opd_tr", 0, type=int),
            "ipd_general_m": form_source.get(f"cbhi_{sr_no}_ipd_general_m", 0, type=int),
            "ipd_general_f": form_source.get(f"cbhi_{sr_no}_ipd_general_f", 0, type=int),
            "ipd_general_tr": form_source.get(f"cbhi_{sr_no}_ipd_general_tr", 0, type=int),
            "ipd_emergency_m": form_source.get(f"cbhi_{sr_no}_ipd_emergency_m", 0, type=int),
            "ipd_emergency_f": form_source.get(f"cbhi_{sr_no}_ipd_emergency_f", 0, type=int),
            "ipd_emergency_tr": form_source.get(f"cbhi_{sr_no}_ipd_emergency_tr", 0, type=int),
            "general_deaths_m": form_source.get(f"cbhi_{sr_no}_general_deaths_m", 0, type=int),
            "general_deaths_f": form_source.get(f"cbhi_{sr_no}_general_deaths_f", 0, type=int),
            "general_deaths_tr": form_source.get(f"cbhi_{sr_no}_general_deaths_tr", 0, type=int),
            "remarks": (form_source.get(f"cbhi_{sr_no}_remarks", "") or "").strip(),
        }
    return payload


def build_ncd_payload(form_source):
    payload = {}
    for sr_no, _, _ in NON_COMMUNICABLE_DISEASES:
        payload[sr_no] = {
            "general_opd_m": form_source.get(f"ncd_{sr_no}_general_opd_m", 0, type=int),
            "general_opd_f": form_source.get(f"ncd_{sr_no}_general_opd_f", 0, type=int),
            "general_opd_tr": form_source.get(f"ncd_{sr_no}_general_opd_tr", 0, type=int),
            "emergency_opd_m": form_source.get(f"ncd_{sr_no}_emergency_opd_m", 0, type=int),
            "emergency_opd_f": form_source.get(f"ncd_{sr_no}_emergency_opd_f", 0, type=int),
            "emergency_opd_tr": form_source.get(f"ncd_{sr_no}_emergency_opd_tr", 0, type=int),
            "ipd_general_m": form_source.get(f"ncd_{sr_no}_ipd_general_m", 0, type=int),
            "ipd_general_f": form_source.get(f"ncd_{sr_no}_ipd_general_f", 0, type=int),
            "ipd_general_tr": form_source.get(f"ncd_{sr_no}_ipd_general_tr", 0, type=int),
            "ipd_emergency_m": form_source.get(f"ncd_{sr_no}_ipd_emergency_m", 0, type=int),
            "ipd_emergency_f": form_source.get(f"ncd_{sr_no}_ipd_emergency_f", 0, type=int),
            "ipd_emergency_tr": form_source.get(f"ncd_{sr_no}_ipd_emergency_tr", 0, type=int),
            "general_deaths_m": form_source.get(f"ncd_{sr_no}_general_deaths_m", 0, type=int),
            "general_deaths_f": form_source.get(f"ncd_{sr_no}_general_deaths_f", 0, type=int),
            "general_deaths_tr": form_source.get(f"ncd_{sr_no}_general_deaths_tr", 0, type=int),
            "remarks": (form_source.get(f"ncd_{sr_no}_remarks", "") or "").strip(),
        }
    return payload


def cbhi_rows(report):
    data = json.loads(report.entries_json or "{}")
    rows = []
    totals = {
        "general_opd_m": 0, "general_opd_f": 0, "general_opd_tr": 0,
        "emergency_opd_m": 0, "emergency_opd_f": 0, "emergency_opd_tr": 0,
        "ipd_general_m": 0, "ipd_general_f": 0, "ipd_general_tr": 0,
        "ipd_emergency_m": 0, "ipd_emergency_f": 0, "ipd_emergency_tr": 0,
        "overall_m": 0, "overall_f": 0, "overall_tr": 0,
        "general_deaths_m": 0, "general_deaths_f": 0, "general_deaths_tr": 0,
    }

    for sr_no, disease_name, icd_code in COMMUNICABLE_DISEASES:
        values = data.get(sr_no, {})
        row = {
            "sr_no": sr_no,
            "disease_name": disease_name,
            "icd_code": icd_code,
            "general_opd_m": int(values.get("general_opd_m", 0) or 0),
            "general_opd_f": int(values.get("general_opd_f", 0) or 0),
            "general_opd_tr": int(values.get("general_opd_tr", 0) or 0),
            "emergency_opd_m": int(values.get("emergency_opd_m", 0) or 0),
            "emergency_opd_f": int(values.get("emergency_opd_f", 0) or 0),
            "emergency_opd_tr": int(values.get("emergency_opd_tr", 0) or 0),
            "ipd_general_m": int(values.get("ipd_general_m", 0) or 0),
            "ipd_general_f": int(values.get("ipd_general_f", 0) or 0),
            "ipd_general_tr": int(values.get("ipd_general_tr", 0) or 0),
            "ipd_emergency_m": int(values.get("ipd_emergency_m", 0) or 0),
            "ipd_emergency_f": int(values.get("ipd_emergency_f", 0) or 0),
            "ipd_emergency_tr": int(values.get("ipd_emergency_tr", 0) or 0),
            "general_deaths_m": int(values.get("general_deaths_m", 0) or 0),
            "general_deaths_f": int(values.get("general_deaths_f", 0) or 0),
            "general_deaths_tr": int(values.get("general_deaths_tr", 0) or 0),
            "remarks": values.get("remarks", "") or "",
        }
        row["general_opd_total"] = row["general_opd_m"] + row["general_opd_f"] + row["general_opd_tr"]
        row["emergency_opd_total"] = row["emergency_opd_m"] + row["emergency_opd_f"] + row["emergency_opd_tr"]
        row["ipd_general_total"] = row["ipd_general_m"] + row["ipd_general_f"] + row["ipd_general_tr"]
        row["ipd_emergency_total"] = row["ipd_emergency_m"] + row["ipd_emergency_f"] + row["ipd_emergency_tr"]
        row["overall_m"] = row["general_opd_m"] + row["emergency_opd_m"]
        row["overall_f"] = row["general_opd_f"] + row["emergency_opd_f"]
        row["overall_tr"] = row["general_opd_tr"] + row["emergency_opd_tr"]
        row["overall_total"] = row["overall_m"] + row["overall_f"] + row["overall_tr"]
        row["general_deaths_total"] = row["general_deaths_m"] + row["general_deaths_f"] + row["general_deaths_tr"]

        for key in totals:
            totals[key] += row[key]

        rows.append(row)

    totals["general_opd_total"] = totals["general_opd_m"] + totals["general_opd_f"] + totals["general_opd_tr"]
    totals["emergency_opd_total"] = totals["emergency_opd_m"] + totals["emergency_opd_f"] + totals["emergency_opd_tr"]
    totals["ipd_general_total"] = totals["ipd_general_m"] + totals["ipd_general_f"] + totals["ipd_general_tr"]
    totals["ipd_emergency_total"] = totals["ipd_emergency_m"] + totals["ipd_emergency_f"] + totals["ipd_emergency_tr"]
    totals["overall_total"] = totals["overall_m"] + totals["overall_f"] + totals["overall_tr"]
    totals["general_deaths_total"] = totals["general_deaths_m"] + totals["general_deaths_f"] + totals["general_deaths_tr"]
    return rows, totals


def ncd_rows(report):
    data = json.loads(report.entries_json or "{}")
    rows = []
    totals = {
        "general_opd_m": 0, "general_opd_f": 0, "general_opd_tr": 0,
        "emergency_opd_m": 0, "emergency_opd_f": 0, "emergency_opd_tr": 0,
        "ipd_general_m": 0, "ipd_general_f": 0, "ipd_general_tr": 0,
        "ipd_emergency_m": 0, "ipd_emergency_f": 0, "ipd_emergency_tr": 0,
        "overall_m": 0, "overall_f": 0, "overall_tr": 0,
        "general_deaths_m": 0, "general_deaths_f": 0, "general_deaths_tr": 0,
    }

    for sr_no, disease_name, icd_code in NON_COMMUNICABLE_DISEASES:
        values = data.get(sr_no, {})
        row = {
            "sr_no": sr_no,
            "disease_name": disease_name,
            "icd_code": icd_code,
            "general_opd_m": int(values.get("general_opd_m", 0) or 0),
            "general_opd_f": int(values.get("general_opd_f", 0) or 0),
            "general_opd_tr": int(values.get("general_opd_tr", 0) or 0),
            "emergency_opd_m": int(values.get("emergency_opd_m", 0) or 0),
            "emergency_opd_f": int(values.get("emergency_opd_f", 0) or 0),
            "emergency_opd_tr": int(values.get("emergency_opd_tr", 0) or 0),
            "ipd_general_m": int(values.get("ipd_general_m", 0) or 0),
            "ipd_general_f": int(values.get("ipd_general_f", 0) or 0),
            "ipd_general_tr": int(values.get("ipd_general_tr", 0) or 0),
            "ipd_emergency_m": int(values.get("ipd_emergency_m", 0) or 0),
            "ipd_emergency_f": int(values.get("ipd_emergency_f", 0) or 0),
            "ipd_emergency_tr": int(values.get("ipd_emergency_tr", 0) or 0),
            "general_deaths_m": int(values.get("general_deaths_m", 0) or 0),
            "general_deaths_f": int(values.get("general_deaths_f", 0) or 0),
            "general_deaths_tr": int(values.get("general_deaths_tr", 0) or 0),
            "remarks": values.get("remarks", "") or "",
        }
        row["general_opd_total"] = row["general_opd_m"] + row["general_opd_f"] + row["general_opd_tr"]
        row["emergency_opd_total"] = row["emergency_opd_m"] + row["emergency_opd_f"] + row["emergency_opd_tr"]
        row["ipd_general_total"] = row["ipd_general_m"] + row["ipd_general_f"] + row["ipd_general_tr"]
        row["ipd_emergency_total"] = row["ipd_emergency_m"] + row["ipd_emergency_f"] + row["ipd_emergency_tr"]
        row["overall_m"] = row["general_opd_m"] + row["emergency_opd_m"]
        row["overall_f"] = row["general_opd_f"] + row["emergency_opd_f"]
        row["overall_tr"] = row["general_opd_tr"] + row["emergency_opd_tr"]
        row["overall_total"] = row["overall_m"] + row["overall_f"] + row["overall_tr"]
        row["general_deaths_total"] = row["general_deaths_m"] + row["general_deaths_f"] + row["general_deaths_tr"]

        for key in totals:
            totals[key] += row[key]

        rows.append(row)

    totals["general_opd_total"] = totals["general_opd_m"] + totals["general_opd_f"] + totals["general_opd_tr"]
    totals["emergency_opd_total"] = totals["emergency_opd_m"] + totals["emergency_opd_f"] + totals["emergency_opd_tr"]
    totals["ipd_general_total"] = totals["ipd_general_m"] + totals["ipd_general_f"] + totals["ipd_general_tr"]
    totals["ipd_emergency_total"] = totals["ipd_emergency_m"] + totals["ipd_emergency_f"] + totals["ipd_emergency_tr"]
    totals["overall_total"] = totals["overall_m"] + totals["overall_f"] + totals["overall_tr"]
    totals["general_deaths_total"] = totals["general_deaths_m"] + totals["general_deaths_f"] + totals["general_deaths_tr"]
    return rows, totals


def create_cbhi_excel(report, rows, totals):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CBHI Form-1"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    title_font = Font(bold=True, size=11)
    bold_font = Font(bold=True, size=9)

    headers = [
        "Sr. No.", "Disease / ICD Code",
        "General OPD M", "General OPD F", "General OPD Tr", "General OPD Total",
        "Emergency OPD M", "Emergency OPD F", "Emergency OPD Tr", "Emergency OPD Total",
        "IPD from General M", "IPD from General F", "IPD from General Tr", "IPD from General Total",
        "IPD from Emergency M", "IPD from Emergency F", "IPD from Emergency Tr", "IPD from Emergency Total",
        "Overall M", "Overall F", "Overall Tr", "Overall Total",
        "Deaths M", "Deaths F", "Deaths Tr", "Deaths Total",
        "Remarks",
    ]

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet.cell(row=1, column=1, value="CENTRAL BUREAU OF HEALTH INTELLIGENCE (CBHI) - ANNEXURE-B (FORM-1)").font = title_font
    sheet.cell(row=1, column=1).alignment = center
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sheet.cell(row=2, column=1, value="MONTHLY REPORT ON INSTITUTIONAL CASES AND DEATHS IN THE REPORTING UNIT DUE TO COMMUNICABLE DISEASES").font = title_font
    sheet.cell(row=2, column=1).alignment = center
    sheet.cell(row=3, column=1, value=f"Month: {report.month}")
    sheet.cell(row=3, column=5, value=f"Year: {report.year}")
    sheet.cell(row=4, column=1, value="Name of Health Establishment")
    sheet.cell(row=4, column=2, value=report.health_establishment_name)
    sheet.cell(row=5, column=1, value="Address / Phone")
    sheet.cell(row=5, column=2, value=report.postal_address_phone)

    header_row = 7
    for index, value in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=index, value=value)
        cell.fill = header_fill
        cell.font = bold_font
        cell.border = border
        cell.alignment = center

    row_index = header_row + 1
    for row in rows:
        values = [
            row["sr_no"], f"{row['disease_name']} ({row['icd_code']})",
            row["general_opd_m"], row["general_opd_f"], row["general_opd_tr"], row["general_opd_total"],
            row["emergency_opd_m"], row["emergency_opd_f"], row["emergency_opd_tr"], row["emergency_opd_total"],
            row["ipd_general_m"], row["ipd_general_f"], row["ipd_general_tr"], row["ipd_general_total"],
            row["ipd_emergency_m"], row["ipd_emergency_f"], row["ipd_emergency_tr"], row["ipd_emergency_total"],
            row["overall_m"], row["overall_f"], row["overall_tr"], row["overall_total"],
            row["general_deaths_m"], row["general_deaths_f"], row["general_deaths_tr"], row["general_deaths_total"],
            row["remarks"],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.border = border
            cell.alignment = left if column in {2, 27} else center
        row_index += 1

    total_values = [
        "", "TOTAL",
        totals["general_opd_m"], totals["general_opd_f"], totals["general_opd_tr"], totals["general_opd_total"],
        totals["emergency_opd_m"], totals["emergency_opd_f"], totals["emergency_opd_tr"], totals["emergency_opd_total"],
        totals["ipd_general_m"], totals["ipd_general_f"], totals["ipd_general_tr"], totals["ipd_general_total"],
        totals["ipd_emergency_m"], totals["ipd_emergency_f"], totals["ipd_emergency_tr"], totals["ipd_emergency_total"],
        totals["overall_m"], totals["overall_f"], totals["overall_tr"], totals["overall_total"],
        totals["general_deaths_m"], totals["general_deaths_f"], totals["general_deaths_tr"], totals["general_deaths_total"],
        "",
    ]
    for column, value in enumerate(total_values, start=1):
        cell = sheet.cell(row=row_index, column=column, value=value)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = left if column == 2 else center

    sheet.column_dimensions["A"].width = 7
    sheet.column_dimensions["B"].width = 36
    for col in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]:
        sheet.column_dimensions[col].width = 10
    sheet.column_dimensions["AA"].width = 12
    sheet.column_dimensions["AB"].width = 20

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def create_ncd_excel(report, rows, totals):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CBHI Form-2"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    title_font = Font(bold=True, size=11)
    bold_font = Font(bold=True, size=9)

    headers = [
        "Sr. No.", "Disease / ICD Code",
        "General OPD M", "General OPD F", "General OPD Tr", "General OPD Total",
        "Emergency OPD M", "Emergency OPD F", "Emergency OPD Tr", "Emergency OPD Total",
        "IPD from General M", "IPD from General F", "IPD from General Tr", "IPD from General Total",
        "IPD from Emergency M", "IPD from Emergency F", "IPD from Emergency Tr", "IPD from Emergency Total",
        "Overall M", "Overall F", "Overall Tr", "Overall Total",
        "Deaths M", "Deaths F", "Deaths Tr", "Deaths Total",
        "Remarks",
    ]

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet.cell(row=1, column=1, value="CENTRAL BUREAU OF HEALTH INTELLIGENCE (CBHI) - ANNEXURE-C (FORM-2)").font = title_font
    sheet.cell(row=1, column=1).alignment = center
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sheet.cell(row=2, column=1, value="MONTHLY REPORT ON INSTITUTIONAL CASES AND DEATHS IN THE REPORTING UNIT DUE TO NON-COMMUNICABLE DISEASES").font = title_font
    sheet.cell(row=2, column=1).alignment = center
    sheet.cell(row=3, column=1, value=f"Month: {report.month}")
    sheet.cell(row=3, column=5, value=f"Year: {report.year}")
    sheet.cell(row=4, column=1, value="Name of Health Establishment")
    sheet.cell(row=4, column=2, value=report.health_establishment_name)
    sheet.cell(row=5, column=1, value="Address / Phone")
    sheet.cell(row=5, column=2, value=report.postal_address_phone)

    header_row = 7
    for index, value in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=index, value=value)
        cell.fill = header_fill
        cell.font = bold_font
        cell.border = border
        cell.alignment = center

    row_index = header_row + 1
    for row in rows:
        values = [
            row["sr_no"], f"{row['disease_name']} ({row['icd_code']})",
            row["general_opd_m"], row["general_opd_f"], row["general_opd_tr"], row["general_opd_total"],
            row["emergency_opd_m"], row["emergency_opd_f"], row["emergency_opd_tr"], row["emergency_opd_total"],
            row["ipd_general_m"], row["ipd_general_f"], row["ipd_general_tr"], row["ipd_general_total"],
            row["ipd_emergency_m"], row["ipd_emergency_f"], row["ipd_emergency_tr"], row["ipd_emergency_total"],
            row["overall_m"], row["overall_f"], row["overall_tr"], row["overall_total"],
            row["general_deaths_m"], row["general_deaths_f"], row["general_deaths_tr"], row["general_deaths_total"],
            row["remarks"],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.border = border
            cell.alignment = left if column in {2, 27} else center
        row_index += 1

    total_values = [
        "", "TOTAL",
        totals["general_opd_m"], totals["general_opd_f"], totals["general_opd_tr"], totals["general_opd_total"],
        totals["emergency_opd_m"], totals["emergency_opd_f"], totals["emergency_opd_tr"], totals["emergency_opd_total"],
        totals["ipd_general_m"], totals["ipd_general_f"], totals["ipd_general_tr"], totals["ipd_general_total"],
        totals["ipd_emergency_m"], totals["ipd_emergency_f"], totals["ipd_emergency_tr"], totals["ipd_emergency_total"],
        totals["overall_m"], totals["overall_f"], totals["overall_tr"], totals["overall_total"],
        totals["general_deaths_m"], totals["general_deaths_f"], totals["general_deaths_tr"], totals["general_deaths_total"],
        "",
    ]
    for column, value in enumerate(total_values, start=1):
        cell = sheet.cell(row=row_index, column=column, value=value)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = left if column == 2 else center

    sheet.column_dimensions["A"].width = 7
    sheet.column_dimensions["B"].width = 36
    for col in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]:
        sheet.column_dimensions[col].width = 10
    sheet.column_dimensions["AA"].width = 12
    sheet.column_dimensions["AB"].width = 20

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_consolidated_rows(cbhi_reports=None, ncd_reports=None, include_owner=False):
    rows = []
    cbhi_reports = cbhi_reports or []
    ncd_reports = ncd_reports or []

    for report in cbhi_reports:
        _, totals = cbhi_rows(report)
        row = {
            "report_type": "CBHI Form-1",
            "institution": report.health_establishment_name,
            "district": "-",
            "month_year": f"{report.month} {report.year}",
            "metric_one_label": "Overall Cases",
            "metric_one": totals["overall_total"],
            "metric_two_label": "Deaths",
            "metric_two": totals["general_deaths_total"],
            "created_at": report.created_at,
            "view_url": url_for("cbhi_report_view", report_id=report.id),
        }
        if include_owner:
            row["owner_username"] = report.cbhi_owner.username
            row["owner_email"] = report.cbhi_owner.email
            row["view_url"] = url_for("admin_cbhi_report_view", report_id=report.id)
        rows.append(row)

    for report in ncd_reports:
        _, totals = ncd_rows(report)
        row = {
            "report_type": "CBHI Form-2",
            "institution": report.health_establishment_name,
            "district": "-",
            "month_year": f"{report.month} {report.year}",
            "metric_one_label": "Overall Cases",
            "metric_one": totals["overall_total"],
            "metric_two_label": "Deaths",
            "metric_two": totals["general_deaths_total"],
            "created_at": report.created_at,
            "view_url": url_for("ncd_report_view", report_id=report.id),
        }
        if include_owner:
            row["owner_username"] = report.ncd_owner.username
            row["owner_email"] = report.ncd_owner.email
            row["view_url"] = url_for("admin_ncd_report_view", report_id=report.id)
        rows.append(row)

    rows.sort(key=lambda item: item["created_at"], reverse=True)
    return rows


CONSOLIDATED_MODULE_LABELS = {
    "all": "All Modules",
    "cbhi1": "CBHI Form-1",
    "cbhi2": "CBHI Form-2",
}


def normalized_module_filter(value):
    module = (value or "all").strip().lower()
    return module if module in CONSOLIDATED_MODULE_LABELS else "all"


def filter_consolidated_rows(rows, module):
    if module == "all":
        return rows
    expected_type = CONSOLIDATED_MODULE_LABELS[module]
    return [row for row in rows if row["report_type"] == expected_type]


def build_consolidated_excel(title, rows, include_owner=False):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Consolidated Report"

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill("solid", fgColor="C2410C")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)

    headers = ["Report Type", "Institution", "District", "Month / Year", "Metric 1", "Value 1", "Metric 2", "Value 2", "Created"]
    if include_owner:
        headers.extend(["Owner Username", "Owner Email"])

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = sheet.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center")

    for index, value in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=index, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_index = 3
    for row in rows:
        values = [
            row["report_type"],
            row["institution"],
            row["district"],
            row["month_year"],
            row["metric_one_label"],
            row["metric_one"],
            row["metric_two_label"],
            row["metric_two"],
            row["created_at"].strftime("%Y-%m-%d"),
        ]
        if include_owner:
            values.extend([row["owner_username"], row["owner_email"]])

        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if column != 2 else "left", vertical="center", wrap_text=True)
        row_index += 1

    for column_index, width in enumerate([18, 30, 18, 16, 14, 12, 14, 12, 14, 18, 26], start=1):
        if column_index <= len(headers):
            sheet.column_dimensions[chr(64 + column_index)].width = width

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def login_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if "user_id" not in session:
            flash("Please login first.", "warning")
            return redirect(url_for("login"))
        user = User.query.get(session["user_id"])
        if not user or not user.is_active:
            session.pop("user_id", None)
            flash("Your account is inactive. Contact admin.", "danger")
            return redirect(url_for("login"))
        return view_func(*args, **kwargs)

    return wrapped


def admin_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if "admin_id" not in session:
            flash("Admin login required.", "warning")
            return redirect(url_for("admin_login"))
        admin = User.query.get(session["admin_id"])
        if not admin or not admin.is_admin or not admin.is_active:
            session.pop("admin_id", None)
            flash("Unauthorized admin access.", "danger")
            return redirect(url_for("admin_login"))
        return view_func(*args, **kwargs)

    return wrapped


def current_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    user = User.query.get(user_id)
    if not user or not user.is_active:
        return None
    return user


def current_admin():
    admin_id = session.get("admin_id")
    if not admin_id:
        return None
    admin = User.query.get(admin_id)
    if admin and admin.is_admin and admin.is_active:
        return admin
    return None


def find_user_by_credential(credential: str):
    """Find user by username, email, or numeric user ID."""
    credential = (credential or "").strip()
    if not credential:
        return None

    user = User.query.filter(func.lower(User.username) == credential.lower()).first()
    if not user and "@" in credential:
        user = User.query.filter(func.lower(User.email) == credential.lower()).first()
    if not user and credential.isdigit():
        user = User.query.get(int(credential))
    return user


@app.route("/")
def index():
    return render_template("index.html", user=current_user())


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        if not all([username, email, password]):
            flash("All fields are required.", "danger")
            return redirect(url_for("register"))

        if len(password) < 8:
            flash("Password must be at least 8 characters.", "danger")
            return redirect(url_for("register"))

        if User.query.filter_by(username=username).first():
            flash("Username already exists.", "danger")
            return redirect(url_for("register"))

        if User.query.filter_by(email=email).first():
            flash("Email already registered.", "danger")
            return redirect(url_for("register"))

        user = User(username=username, email=email)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()

        flash("Registration successful. Please login.", "success")
        return redirect(url_for("login"))

    return render_template("register.html", user=current_user())


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        credential = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        user = find_user_by_credential(credential)
        if user and user.check_password(password):
            if not user.is_active:
                flash("Your account is inactive. Contact admin.", "danger")
                return redirect(url_for("login"))
            session["user_id"] = user.id
            flash(f"Welcome back, {user.username}.", "success")
            return redirect(url_for("dashboard"))

        flash("Invalid username/email/user ID or password.", "danger")
        return redirect(url_for("login"))

    return render_template("login.html", user=current_user())


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip().lower()
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not all([username, email, new_password, confirm_password]):
            flash("All fields are required.", "danger")
            return redirect(url_for("forgot_password"))

        if new_password != confirm_password:
            flash("Passwords do not match.", "danger")
            return redirect(url_for("forgot_password"))

        if len(new_password) < 8:
            flash("Password must be at least 8 characters.", "danger")
            return redirect(url_for("forgot_password"))

        user = User.query.filter_by(username=username, email=email).first()
        if not user:
            flash("No account matched that username and email.", "danger")
            return redirect(url_for("forgot_password"))

        user.set_password(new_password)
        db.session.commit()
        flash("Password updated successfully. Please login.", "success")
        return redirect(url_for("login"))

    return render_template("forgot_password.html", user=current_user())


@app.route("/dashboard")
@login_required
def dashboard():
    user = current_user()
    cbhi_reports = CbhiReport.query.filter_by(user_id=user.id).order_by(CbhiReport.created_at.desc()).all()
    ncd_reports = NcdReport.query.filter_by(user_id=user.id).order_by(NcdReport.created_at.desc()).all()

    cbhi_summaries = []
    for report in cbhi_reports[:5]:
        _, totals = cbhi_rows(report)
        cbhi_summaries.append({
            "report": report,
            "overall_total": totals["overall_total"],
            "deaths_total": totals["general_deaths_total"],
        })

    ncd_summaries = []
    for report in ncd_reports[:5]:
        _, totals = ncd_rows(report)
        ncd_summaries.append({
            "report": report,
            "overall_total": totals["overall_total"],
            "deaths_total": totals["general_deaths_total"],
        })

    return render_template(
        "dashboard.html",
        user=user,
        cbhi_reports=cbhi_reports,
        cbhi_summaries=cbhi_summaries,
        ncd_reports=ncd_reports,
        ncd_summaries=ncd_summaries,
    )


@app.route("/logout")
def logout():
    session.clear()
    flash("Logged out successfully.", "info")
    return redirect(url_for("index"))


@app.route("/cbhi/report/new", methods=["GET", "POST"])
@login_required
def new_cbhi_report():
    user = current_user()
    disease_ids = [sr_no for sr_no, _, _ in COMMUNICABLE_DISEASES]
    if request.method == "POST":
        health_establishment_name = request.form.get("health_establishment_name", "").strip()
        postal_address_phone = request.form.get("postal_address_phone", "").strip()
        month = request.form.get("month", "").strip()
        year = request.form.get("year", "").strip()

        if not all([health_establishment_name, postal_address_phone, month, year]):
            flash("Health establishment, address, month, and year are required.", "danger")
            return redirect(url_for("new_cbhi_report"))

        report = CbhiReport(
            user_id=user.id,
            health_establishment_name=health_establishment_name,
            postal_address_phone=postal_address_phone,
            month=month,
            year=year,
            entries_json=json.dumps(build_cbhi_payload(request.form)),
            approving_authority_name=request.form.get("approving_authority_name", "").strip(),
            approving_authority_designation=request.form.get("approving_authority_designation", "").strip(),
            official_email=request.form.get("official_email", "").strip(),
            official_phone=request.form.get("official_phone", "").strip(),
        )
        db.session.add(report)
        db.session.commit()
        flash("CBHI Form-1 report submitted.", "success")
        return redirect(url_for("cbhi_reports"))

    return render_template("cbhi_report_form.html", user=user, diseases=COMMUNICABLE_DISEASES, disease_ids=disease_ids)


@app.route("/cbhi/reports")
@login_required
def cbhi_reports():
    user = current_user()
    reports = CbhiReport.query.filter_by(user_id=user.id).order_by(CbhiReport.created_at.desc()).all()
    return render_template("cbhi_reports.html", user=user, reports=reports)


@app.route("/cbhi/report/<int:report_id>")
@login_required
def cbhi_report_view(report_id):
    user = current_user()
    report = CbhiReport.query.get_or_404(report_id)
    if report.user_id != user.id:
        flash("You do not have permission to view this report.", "danger")
        return redirect(url_for("cbhi_reports"))
    rows, totals = cbhi_rows(report)
    return render_template("cbhi_report_view.html", user=user, report=report, rows=rows, totals=totals)


@app.route("/cbhi/report/<int:report_id>/edit", methods=["GET", "POST"])
@login_required
def cbhi_report_edit(report_id):
    user = current_user()
    report = CbhiReport.query.get_or_404(report_id)
    if report.user_id != user.id:
        flash("You do not have permission to edit this report.", "danger")
        return redirect(url_for("cbhi_reports"))

    if request.method == "POST":
        report.health_establishment_name = request.form.get("health_establishment_name", "").strip()
        report.postal_address_phone = request.form.get("postal_address_phone", "").strip()
        report.month = request.form.get("month", "").strip()
        report.year = request.form.get("year", "").strip()
        if not all([report.health_establishment_name, report.postal_address_phone, report.month, report.year]):
            flash("Health establishment, address, month, and year are required.", "danger")
            return redirect(url_for("cbhi_report_edit", report_id=report.id))
        report.entries_json = json.dumps(build_cbhi_payload(request.form))
        report.approving_authority_name = request.form.get("approving_authority_name", "").strip()
        report.approving_authority_designation = request.form.get("approving_authority_designation", "").strip()
        report.official_email = request.form.get("official_email", "").strip()
        report.official_phone = request.form.get("official_phone", "").strip()
        db.session.commit()
        flash("CBHI Form-1 report updated.", "success")
        return redirect(url_for("cbhi_report_view", report_id=report.id))

    return render_template(
        "cbhi_report_form.html",
        user=user,
        diseases=COMMUNICABLE_DISEASES,
        disease_ids=[sr_no for sr_no, _, _ in COMMUNICABLE_DISEASES],
        report=report,
        form_action=url_for("cbhi_report_edit", report_id=report.id),
        submit_label="Save CBHI Form-1 Report",
        cancel_url=url_for("cbhi_report_view", report_id=report.id),
        initial_payload=json.loads(report.entries_json or "{}"),
    )


@app.route("/cbhi/report/<int:report_id>/print")
@login_required
def cbhi_report_print(report_id):
    user = current_user()
    report = CbhiReport.query.get_or_404(report_id)
    if report.user_id != user.id:
        flash("You do not have permission.", "danger")
        return redirect(url_for("cbhi_reports"))
    rows, totals = cbhi_rows(report)
    return render_template("cbhi_report_print.html", report=report, rows=rows, totals=totals)


@app.route("/cbhi/report/<int:report_id>/export/csv")
@login_required
def cbhi_report_csv(report_id):
    user = current_user()
    report = CbhiReport.query.get_or_404(report_id)
    if report.user_id != user.id:
        flash("You do not have permission.", "danger")
        return redirect(url_for("cbhi_reports"))
    rows, totals = cbhi_rows(report)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["CENTRAL BUREAU OF HEALTH INTELLIGENCE (CBHI)", "ANNEXURE-B", "FORM-1"])
    writer.writerow(["MONTHLY REPORT ON INSTITUTIONAL CASES AND DEATHS IN THE REPORTING UNIT DUE TO COMMUNICABLE DISEASES"])
    writer.writerow([f"Month: {report.month}", f"Year: {report.year}"])
    writer.writerow([f"Name of Health Establishment: {report.health_establishment_name}"])
    writer.writerow([f"Address / Phone: {report.postal_address_phone}"])
    writer.writerow([
        "Sr. No.", "Disease / ICD Code",
        "General OPD M", "General OPD F", "General OPD Tr", "General OPD Total",
        "Emergency OPD M", "Emergency OPD F", "Emergency OPD Tr", "Emergency OPD Total",
        "IPD from General M", "IPD from General F", "IPD from General Tr", "IPD from General Total",
        "IPD from Emergency M", "IPD from Emergency F", "IPD from Emergency Tr", "IPD from Emergency Total",
        "Overall M", "Overall F", "Overall Tr", "Overall Total",
        "Deaths M", "Deaths F", "Deaths Tr", "Deaths Total", "Remarks",
    ])
    for row in rows:
        writer.writerow([
            row["sr_no"], f"{row['disease_name']} ({row['icd_code']})",
            row["general_opd_m"], row["general_opd_f"], row["general_opd_tr"], row["general_opd_total"],
            row["emergency_opd_m"], row["emergency_opd_f"], row["emergency_opd_tr"], row["emergency_opd_total"],
            row["ipd_general_m"], row["ipd_general_f"], row["ipd_general_tr"], row["ipd_general_total"],
            row["ipd_emergency_m"], row["ipd_emergency_f"], row["ipd_emergency_tr"], row["ipd_emergency_total"],
            row["overall_m"], row["overall_f"], row["overall_tr"], row["overall_total"],
            row["general_deaths_m"], row["general_deaths_f"], row["general_deaths_tr"], row["general_deaths_total"],
            row["remarks"],
        ])
    writer.writerow([
        "", "TOTAL",
        totals["general_opd_m"], totals["general_opd_f"], totals["general_opd_tr"], totals["general_opd_total"],
        totals["emergency_opd_m"], totals["emergency_opd_f"], totals["emergency_opd_tr"], totals["emergency_opd_total"],
        totals["ipd_general_m"], totals["ipd_general_f"], totals["ipd_general_tr"], totals["ipd_general_total"],
        totals["ipd_emergency_m"], totals["ipd_emergency_f"], totals["ipd_emergency_tr"], totals["ipd_emergency_total"],
        totals["overall_m"], totals["overall_f"], totals["overall_tr"], totals["overall_total"],
        totals["general_deaths_m"], totals["general_deaths_f"], totals["general_deaths_tr"], totals["general_deaths_total"],
        "",
    ])
    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    response.headers["Content-Disposition"] = f"attachment; filename={report.health_establishment_name}_{report.month}_{report.year}_cbhi_form1.csv".replace(" ", "_")
    return response


@app.route("/cbhi/report/<int:report_id>/export/excel")
@login_required
def cbhi_report_excel(report_id):
    user = current_user()
    report = CbhiReport.query.get_or_404(report_id)
    if report.user_id != user.id:
        flash("You do not have permission.", "danger")
        return redirect(url_for("cbhi_reports"))
    rows, totals = cbhi_rows(report)
    output = create_cbhi_excel(report, rows, totals)
    response = make_response(output.read())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = f"attachment; filename={report.health_establishment_name}_{report.month}_{report.year}_cbhi_form1.xlsx".replace(" ", "_")
    return response


@app.route("/ncd/report/new", methods=["GET", "POST"])
@login_required
def new_ncd_report():
    user = current_user()
    disease_ids = [sr_no for sr_no, _, _ in NON_COMMUNICABLE_DISEASES]
    if request.method == "POST":
        health_establishment_name = request.form.get("health_establishment_name", "").strip()
        postal_address_phone = request.form.get("postal_address_phone", "").strip()
        month = request.form.get("month", "").strip()
        year = request.form.get("year", "").strip()

        if not all([health_establishment_name, postal_address_phone, month, year]):
            flash("Health establishment, address, month, and year are required.", "danger")
            return redirect(url_for("new_ncd_report"))

        report = NcdReport(
            user_id=user.id,
            health_establishment_name=health_establishment_name,
            postal_address_phone=postal_address_phone,
            month=month,
            year=year,
            entries_json=json.dumps(build_ncd_payload(request.form)),
            approving_authority_name=request.form.get("approving_authority_name", "").strip(),
            approving_authority_designation=request.form.get("approving_authority_designation", "").strip(),
            official_email=request.form.get("official_email", "").strip(),
            official_phone=request.form.get("official_phone", "").strip(),
        )
        db.session.add(report)
        db.session.commit()
        flash("CBHI Form-2 report submitted.", "success")
        return redirect(url_for("ncd_reports"))

    return render_template("ncd_report_form.html", user=user, diseases=NON_COMMUNICABLE_DISEASES, disease_ids=disease_ids)


@app.route("/ncd/reports")
@login_required
def ncd_reports():
    user = current_user()
    reports = NcdReport.query.filter_by(user_id=user.id).order_by(NcdReport.created_at.desc()).all()
    return render_template("ncd_reports.html", user=user, reports=reports)


@app.route("/ncd/report/<int:report_id>")
@login_required
def ncd_report_view(report_id):
    user = current_user()
    report = NcdReport.query.get_or_404(report_id)
    if report.user_id != user.id:
        flash("You do not have permission to view this report.", "danger")
        return redirect(url_for("ncd_reports"))
    rows, totals = ncd_rows(report)
    return render_template("ncd_report_view.html", user=user, report=report, rows=rows, totals=totals)


@app.route("/ncd/report/<int:report_id>/edit", methods=["GET", "POST"])
@login_required
def ncd_report_edit(report_id):
    user = current_user()
    report = NcdReport.query.get_or_404(report_id)
    if report.user_id != user.id:
        flash("You do not have permission to edit this report.", "danger")
        return redirect(url_for("ncd_reports"))

    if request.method == "POST":
        report.health_establishment_name = request.form.get("health_establishment_name", "").strip()
        report.postal_address_phone = request.form.get("postal_address_phone", "").strip()
        report.month = request.form.get("month", "").strip()
        report.year = request.form.get("year", "").strip()
        if not all([report.health_establishment_name, report.postal_address_phone, report.month, report.year]):
            flash("Health establishment, address, month, and year are required.", "danger")
            return redirect(url_for("ncd_report_edit", report_id=report.id))
        report.entries_json = json.dumps(build_ncd_payload(request.form))
        report.approving_authority_name = request.form.get("approving_authority_name", "").strip()
        report.approving_authority_designation = request.form.get("approving_authority_designation", "").strip()
        report.official_email = request.form.get("official_email", "").strip()
        report.official_phone = request.form.get("official_phone", "").strip()
        db.session.commit()
        flash("CBHI Form-2 report updated.", "success")
        return redirect(url_for("ncd_report_view", report_id=report.id))

    return render_template(
        "ncd_report_form.html",
        user=user,
        diseases=NON_COMMUNICABLE_DISEASES,
        disease_ids=[sr_no for sr_no, _, _ in NON_COMMUNICABLE_DISEASES],
        report=report,
        form_action=url_for("ncd_report_edit", report_id=report.id),
        submit_label="Save CBHI Form-2 Report",
        cancel_url=url_for("ncd_report_view", report_id=report.id),
        initial_payload=json.loads(report.entries_json or "{}"),
    )


@app.route("/ncd/report/<int:report_id>/print")
@login_required
def ncd_report_print(report_id):
    user = current_user()
    report = NcdReport.query.get_or_404(report_id)
    if report.user_id != user.id:
        flash("You do not have permission.", "danger")
        return redirect(url_for("ncd_reports"))
    rows, totals = ncd_rows(report)
    return render_template("ncd_report_print.html", report=report, rows=rows, totals=totals)


@app.route("/ncd/report/<int:report_id>/export/csv")
@login_required
def ncd_report_csv(report_id):
    user = current_user()
    report = NcdReport.query.get_or_404(report_id)
    if report.user_id != user.id:
        flash("You do not have permission.", "danger")
        return redirect(url_for("ncd_reports"))
    rows, totals = ncd_rows(report)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["CENTRAL BUREAU OF HEALTH INTELLIGENCE (CBHI)", "ANNEXURE-C", "FORM-2"])
    writer.writerow(["MONTHLY REPORT ON INSTITUTIONAL CASES AND DEATHS IN THE REPORTING UNIT DUE TO NON-COMMUNICABLE DISEASES"])
    writer.writerow([f"Month: {report.month}", f"Year: {report.year}"])
    writer.writerow([f"Name of Health Establishment: {report.health_establishment_name}"])
    writer.writerow([f"Address / Phone: {report.postal_address_phone}"])
    writer.writerow([
        "Sr. No.", "Disease / ICD Code",
        "General OPD M", "General OPD F", "General OPD Tr", "General OPD Total",
        "Emergency OPD M", "Emergency OPD F", "Emergency OPD Tr", "Emergency OPD Total",
        "IPD from General M", "IPD from General F", "IPD from General Tr", "IPD from General Total",
        "IPD from Emergency M", "IPD from Emergency F", "IPD from Emergency Tr", "IPD from Emergency Total",
        "Overall M", "Overall F", "Overall Tr", "Overall Total",
        "Deaths M", "Deaths F", "Deaths Tr", "Deaths Total", "Remarks",
    ])
    for row in rows:
        writer.writerow([
            row["sr_no"], f"{row['disease_name']} ({row['icd_code']})",
            row["general_opd_m"], row["general_opd_f"], row["general_opd_tr"], row["general_opd_total"],
            row["emergency_opd_m"], row["emergency_opd_f"], row["emergency_opd_tr"], row["emergency_opd_total"],
            row["ipd_general_m"], row["ipd_general_f"], row["ipd_general_tr"], row["ipd_general_total"],
            row["ipd_emergency_m"], row["ipd_emergency_f"], row["ipd_emergency_tr"], row["ipd_emergency_total"],
            row["overall_m"], row["overall_f"], row["overall_tr"], row["overall_total"],
            row["general_deaths_m"], row["general_deaths_f"], row["general_deaths_tr"], row["general_deaths_total"],
            row["remarks"],
        ])
    writer.writerow([
        "", "TOTAL",
        totals["general_opd_m"], totals["general_opd_f"], totals["general_opd_tr"], totals["general_opd_total"],
        totals["emergency_opd_m"], totals["emergency_opd_f"], totals["emergency_opd_tr"], totals["emergency_opd_total"],
        totals["ipd_general_m"], totals["ipd_general_f"], totals["ipd_general_tr"], totals["ipd_general_total"],
        totals["ipd_emergency_m"], totals["ipd_emergency_f"], totals["ipd_emergency_tr"], totals["ipd_emergency_total"],
        totals["overall_m"], totals["overall_f"], totals["overall_tr"], totals["overall_total"],
        totals["general_deaths_m"], totals["general_deaths_f"], totals["general_deaths_tr"], totals["general_deaths_total"],
        "",
    ])
    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    response.headers["Content-Disposition"] = f"attachment; filename={report.health_establishment_name}_{report.month}_{report.year}_cbhi_form2.csv".replace(" ", "_")
    return response


@app.route("/ncd/report/<int:report_id>/export/excel")
@login_required
def ncd_report_excel(report_id):
    user = current_user()
    report = NcdReport.query.get_or_404(report_id)
    if report.user_id != user.id:
        flash("You do not have permission.", "danger")
        return redirect(url_for("ncd_reports"))
    rows, totals = ncd_rows(report)
    output = create_ncd_excel(report, rows, totals)
    response = make_response(output.read())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = f"attachment; filename={report.health_establishment_name}_{report.month}_{report.year}_cbhi_form2.xlsx".replace(" ", "_")
    return response


@app.route("/reports/consolidated")
@login_required
def consolidated_reports():
    user = current_user()
    module = normalized_module_filter(request.args.get("module"))
    cbhi_reports_list = CbhiReport.query.filter_by(user_id=user.id).order_by(CbhiReport.created_at.desc()).all()
    ncd_reports_list = NcdReport.query.filter_by(user_id=user.id).order_by(NcdReport.created_at.desc()).all()
    rows = build_consolidated_rows(cbhi_reports_list, ncd_reports_list)
    filtered_rows = filter_consolidated_rows(rows, module)
    return render_template(
        "consolidated_reports.html",
        user=user,
        rows=filtered_rows,
        show_owner=False,
        title="My Consolidated Reports",
        selected_module=module,
        module_options=CONSOLIDATED_MODULE_LABELS,
    )


@app.route("/reports/consolidated/print")
@login_required
def consolidated_reports_print():
    user = current_user()
    module = normalized_module_filter(request.args.get("module"))
    cbhi_reports_list = CbhiReport.query.filter_by(user_id=user.id).order_by(CbhiReport.created_at.desc()).all()
    ncd_reports_list = NcdReport.query.filter_by(user_id=user.id).order_by(NcdReport.created_at.desc()).all()
    rows = build_consolidated_rows(cbhi_reports_list, ncd_reports_list)
    filtered_rows = filter_consolidated_rows(rows, module)
    return render_template(
        "consolidated_reports_print.html",
        rows=filtered_rows,
        show_owner=False,
        title="My Consolidated Reports",
        selected_module=module,
        selected_module_label=CONSOLIDATED_MODULE_LABELS[module],
    )


@app.route("/reports/consolidated/export/csv")
@login_required
def consolidated_reports_csv():
    user = current_user()
    module = normalized_module_filter(request.args.get("module"))
    cbhi_reports_list = CbhiReport.query.filter_by(user_id=user.id).order_by(CbhiReport.created_at.desc()).all()
    ncd_reports_list = NcdReport.query.filter_by(user_id=user.id).order_by(NcdReport.created_at.desc()).all()
    rows = build_consolidated_rows(cbhi_reports_list, ncd_reports_list)
    filtered_rows = filter_consolidated_rows(rows, module)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["report_type", "institution", "district", "month_year", "metric_one_label", "metric_one", "metric_two_label", "metric_two", "created_at"])
    for row in filtered_rows:
        writer.writerow([
            row["report_type"], row["institution"], row["district"], row["month_year"],
            row["metric_one_label"], row["metric_one"], row["metric_two_label"], row["metric_two"],
            row["created_at"].isoformat(),
        ])

    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    filename = "consolidated_reports.csv" if module == "all" else f"consolidated_reports_{module}.csv"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@app.route("/reports/consolidated/export/excel")
@login_required
def consolidated_reports_excel():
    user = current_user()
    module = normalized_module_filter(request.args.get("module"))
    cbhi_reports_list = CbhiReport.query.filter_by(user_id=user.id).order_by(CbhiReport.created_at.desc()).all()
    ncd_reports_list = NcdReport.query.filter_by(user_id=user.id).order_by(NcdReport.created_at.desc()).all()
    rows = build_consolidated_rows(cbhi_reports_list, ncd_reports_list)
    filtered_rows = filter_consolidated_rows(rows, module)
    output = build_consolidated_excel(f"My Consolidated Reports - {CONSOLIDATED_MODULE_LABELS[module]}", filtered_rows)

    response = make_response(output.read())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    filename = "consolidated_reports.xlsx" if module == "all" else f"consolidated_reports_{module}.xlsx"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        credential = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        admin = None
        if credential:
            admin = User.query.filter(func.lower(User.username) == credential.lower()).first()
            if not admin and "@" in credential:
                admin = User.query.filter(func.lower(User.email) == credential.lower()).first()
        if admin and admin.is_admin and admin.check_password(password):
            if not admin.is_active:
                flash("Admin account is inactive.", "danger")
                return redirect(url_for("admin_login"))
            session["admin_id"] = admin.id
            flash(f"Welcome Admin, {admin.username}.", "success")
            return redirect(url_for("admin_dashboard"))

        flash("Invalid admin credentials.", "danger")
        return redirect(url_for("admin_login"))

    return render_template("admin_login.html", user=current_user())


@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    admin = current_admin()
    users = User.query.order_by(User.created_at.desc()).all()
    cbhi_reports = CbhiReport.query.order_by(CbhiReport.created_at.desc()).all()
    ncd_reports = NcdReport.query.order_by(NcdReport.created_at.desc()).all()
    stats = {
        "total_users": len(users),
        "total_reports": len(cbhi_reports) + len(ncd_reports),
        "total_cbhi_reports": len(cbhi_reports),
        "total_ncd_reports": len(ncd_reports),
    }
    cbhi_summaries = []
    for report in cbhi_reports[:10]:
        _, totals = cbhi_rows(report)
        cbhi_summaries.append({
            "report": report,
            "overall_total": totals["overall_total"],
            "deaths_total": totals["general_deaths_total"],
        })
    ncd_summaries = []
    for report in ncd_reports[:10]:
        _, totals = ncd_rows(report)
        ncd_summaries.append({
            "report": report,
            "overall_total": totals["overall_total"],
            "deaths_total": totals["general_deaths_total"],
        })
    return render_template(
        "admin_dashboard.html",
        user=current_user(),
        admin=admin,
        users=users,
        cbhi_reports=cbhi_reports,
        ncd_reports=ncd_reports,
        cbhi_summaries=cbhi_summaries,
        ncd_summaries=ncd_summaries,
        stats=stats,
    )


@app.route("/admin/users")
@admin_required
def admin_users():
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template("admin_users.html", user=current_user(), admin=current_admin(), users=users)


@app.route("/admin/users/new", methods=["GET", "POST"])
@admin_required
def admin_user_create():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        is_admin = request.form.get("is_admin") == "on"

        if not all([username, email, password]):
            flash("Username, email, and password are required.", "danger")
            return redirect(url_for("admin_user_create"))

        if len(password) < 8:
            flash("Password must be at least 8 characters.", "danger")
            return redirect(url_for("admin_user_create"))

        if User.query.filter_by(username=username).first():
            flash("Username already exists.", "danger")
            return redirect(url_for("admin_user_create"))

        if User.query.filter_by(email=email).first():
            flash("Email already exists.", "danger")
            return redirect(url_for("admin_user_create"))

        new_user = User(username=username, email=email, is_admin=is_admin, is_active=True)
        new_user.set_password(password)
        db.session.add(new_user)
        db.session.commit()

        flash(f"Login ID created for {username}.", "success")
        return redirect(url_for("admin_users"))

    return render_template("admin_user_create.html", user=current_user(), admin=current_admin())


@app.route("/admin/users/import", methods=["GET", "POST"])
@admin_required
def admin_users_import():
    if request.method == "POST":
        csv_file = request.files.get("csv_file")
        if not csv_file or not csv_file.filename:
            flash("Please choose a CSV file.", "danger")
            return redirect(url_for("admin_users_import"))

        try:
            csv_text = csv_file.stream.read().decode("utf-8-sig")
        except Exception:
            flash("Could not read CSV file. Use UTF-8 encoding.", "danger")
            return redirect(url_for("admin_users_import"))

        reader = csv.DictReader(io.StringIO(csv_text))
        required_columns = {"username", "email", "password"}
        if not reader.fieldnames:
            flash("CSV is empty or invalid.", "danger")
            return redirect(url_for("admin_users_import"))

        headers = {h.strip().lower() for h in reader.fieldnames if h}
        missing_columns = required_columns - headers
        if missing_columns:
            missing_list = ", ".join(sorted(missing_columns))
            flash(f"Missing required CSV columns: {missing_list}", "danger")
            return redirect(url_for("admin_users_import"))

        created = 0
        skipped = 0
        batch_usernames = set()
        batch_emails = set()

        for row in reader:
            username = (row.get("username") or "").strip()
            email = (row.get("email") or "").strip().lower()
            password = row.get("password") or ""
            is_admin_raw = (row.get("is_admin") or "").strip().lower()
            is_admin = is_admin_raw in {"1", "true", "yes", "y", "on"}

            if not all([username, email, password]):
                skipped += 1
                continue

            if len(password) < 8:
                skipped += 1
                continue

            if username in batch_usernames or email in batch_emails:
                skipped += 1
                continue

            if User.query.filter_by(username=username).first() or User.query.filter_by(email=email).first():
                skipped += 1
                continue

            new_user = User(username=username, email=email, is_admin=is_admin, is_active=True)
            new_user.set_password(password)
            db.session.add(new_user)
            batch_usernames.add(username)
            batch_emails.add(email)
            created += 1

        db.session.commit()
        flash(f"CSV import complete. Created: {created}, Skipped: {skipped}.", "success")
        return redirect(url_for("admin_users"))

    return render_template("admin_user_import.html", user=current_user(), admin=current_admin())


@app.route("/admin/users/import/template.csv")
@admin_required
def admin_users_import_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["username", "email", "password", "is_admin"])
    writer.writerow(["staff1", "staff1@example.com", "StrongPass123", "false"])
    writer.writerow(["manager1", "manager1@example.com", "StrongPass123", "true"])

    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    response.headers["Content-Disposition"] = "attachment; filename=users_import_template.csv"
    return response


@app.route("/admin/user/<int:user_id>/reset-password", methods=["GET", "POST"])
@admin_required
def admin_user_reset_password(user_id):
    target_user = User.query.get_or_404(user_id)

    if request.method == "POST":
        new_password = request.form.get("password", "")
        if len(new_password) < 8:
            flash("Password must be at least 8 characters.", "danger")
            return redirect(url_for("admin_user_reset_password", user_id=target_user.id))

        target_user.set_password(new_password)
        db.session.commit()
        flash(f"Password reset for {target_user.username}.", "success")
        return redirect(url_for("admin_users"))

    return render_template("admin_user_reset_password.html", user=current_user(), admin=current_admin(), target_user=target_user)


@app.route("/admin/user/<int:user_id>/toggle-active", methods=["POST"])
@admin_required
def admin_user_toggle_active(user_id):
    target_user = User.query.get_or_404(user_id)
    admin = current_admin()

    if target_user.id == admin.id and target_user.is_active:
        flash("You cannot deactivate your own account.", "danger")
        return redirect(url_for("admin_users"))

    if target_user.is_admin and target_user.is_active:
        active_admin_count = User.query.filter_by(is_admin=True, is_active=True).count()
        if active_admin_count <= 1:
            flash("Cannot deactivate the last active admin.", "danger")
            return redirect(url_for("admin_users"))

    target_user.is_active = not target_user.is_active
    db.session.commit()
    state = "activated" if target_user.is_active else "deactivated"
    flash(f"{target_user.username} has been {state}.", "info")
    return redirect(url_for("admin_users"))


@app.route("/admin/user/<int:user_id>/delete", methods=["POST"])
@admin_required
def admin_user_delete(user_id):
    target_user = User.query.get_or_404(user_id)
    admin = current_admin()

    if target_user.id == admin.id:
        flash("You cannot delete your own account.", "danger")
        return redirect(url_for("admin_users"))

    if target_user.is_admin:
        total_admin_count = User.query.filter_by(is_admin=True).count()
        if total_admin_count <= 1:
            flash("Cannot delete the last admin account.", "danger")
            return redirect(url_for("admin_users"))

    db.session.delete(target_user)
    db.session.commit()
    flash("User deleted successfully.", "info")
    return redirect(url_for("admin_users"))


@app.route("/admin/cbhi-report/<int:report_id>")
@admin_required
def admin_cbhi_report_view(report_id):
    report = CbhiReport.query.get_or_404(report_id)
    rows, totals = cbhi_rows(report)
    return render_template("admin_cbhi_report_view.html", user=current_user(), admin=current_admin(), report=report, rows=rows, totals=totals)


@app.route("/admin/cbhi-report/<int:report_id>/edit", methods=["GET", "POST"])
@admin_required
def admin_cbhi_report_edit(report_id):
    report = CbhiReport.query.get_or_404(report_id)

    if request.method == "POST":
        report.health_establishment_name = request.form.get("health_establishment_name", "").strip()
        report.postal_address_phone = request.form.get("postal_address_phone", "").strip()
        report.month = request.form.get("month", "").strip()
        report.year = request.form.get("year", "").strip()
        if not all([report.health_establishment_name, report.postal_address_phone, report.month, report.year]):
            flash("Health establishment, address, month, and year are required.", "danger")
            return redirect(url_for("admin_cbhi_report_edit", report_id=report.id))
        report.entries_json = json.dumps(build_cbhi_payload(request.form))
        report.approving_authority_name = request.form.get("approving_authority_name", "").strip()
        report.approving_authority_designation = request.form.get("approving_authority_designation", "").strip()
        report.official_email = request.form.get("official_email", "").strip()
        report.official_phone = request.form.get("official_phone", "").strip()
        db.session.commit()
        flash("CBHI Form-1 report updated.", "success")
        return redirect(url_for("admin_cbhi_report_view", report_id=report.id))

    return render_template(
        "cbhi_report_form.html",
        user=current_user(),
        admin=current_admin(),
        diseases=COMMUNICABLE_DISEASES,
        disease_ids=[sr_no for sr_no, _, _ in COMMUNICABLE_DISEASES],
        report=report,
        form_action=url_for("admin_cbhi_report_edit", report_id=report.id),
        submit_label="Save CBHI Form-1 Report",
        cancel_url=url_for("admin_cbhi_report_view", report_id=report.id),
        initial_payload=json.loads(report.entries_json or "{}"),
    )


@app.route("/admin/cbhi-report/<int:report_id>/delete", methods=["POST"])
@admin_required
def admin_cbhi_report_delete(report_id):
    report = CbhiReport.query.get_or_404(report_id)
    db.session.delete(report)
    db.session.commit()
    flash("CBHI Form-1 report deleted.", "info")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/cbhi-reports/export.csv")
@admin_required
def admin_cbhi_reports_export():
    reports = CbhiReport.query.order_by(CbhiReport.created_at.desc()).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["id", "health_establishment_name", "month", "year", "owner_username", "owner_email", "overall_total", "deaths_total", "created_at"])
    for report in reports:
        _, totals = cbhi_rows(report)
        writer.writerow([
            report.id,
            report.health_establishment_name,
            report.month,
            report.year,
            report.cbhi_owner.username,
            report.cbhi_owner.email,
            totals["overall_total"],
            totals["general_deaths_total"],
            report.created_at.isoformat(),
        ])
    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    response.headers["Content-Disposition"] = "attachment; filename=cbhi_reports.csv"
    return response


@app.route("/admin/ncd-report/<int:report_id>")
@admin_required
def admin_ncd_report_view(report_id):
    report = NcdReport.query.get_or_404(report_id)
    rows, totals = ncd_rows(report)
    return render_template("admin_ncd_report_view.html", user=current_user(), admin=current_admin(), report=report, rows=rows, totals=totals)


@app.route("/admin/ncd-report/<int:report_id>/edit", methods=["GET", "POST"])
@admin_required
def admin_ncd_report_edit(report_id):
    report = NcdReport.query.get_or_404(report_id)

    if request.method == "POST":
        report.health_establishment_name = request.form.get("health_establishment_name", "").strip()
        report.postal_address_phone = request.form.get("postal_address_phone", "").strip()
        report.month = request.form.get("month", "").strip()
        report.year = request.form.get("year", "").strip()
        if not all([report.health_establishment_name, report.postal_address_phone, report.month, report.year]):
            flash("Health establishment, address, month, and year are required.", "danger")
            return redirect(url_for("admin_ncd_report_edit", report_id=report.id))
        report.entries_json = json.dumps(build_ncd_payload(request.form))
        report.approving_authority_name = request.form.get("approving_authority_name", "").strip()
        report.approving_authority_designation = request.form.get("approving_authority_designation", "").strip()
        report.official_email = request.form.get("official_email", "").strip()
        report.official_phone = request.form.get("official_phone", "").strip()
        db.session.commit()
        flash("CBHI Form-2 report updated.", "success")
        return redirect(url_for("admin_ncd_report_view", report_id=report.id))

    return render_template(
        "ncd_report_form.html",
        user=current_user(),
        admin=current_admin(),
        diseases=NON_COMMUNICABLE_DISEASES,
        disease_ids=[sr_no for sr_no, _, _ in NON_COMMUNICABLE_DISEASES],
        report=report,
        form_action=url_for("admin_ncd_report_edit", report_id=report.id),
        submit_label="Save CBHI Form-2 Report",
        cancel_url=url_for("admin_ncd_report_view", report_id=report.id),
        initial_payload=json.loads(report.entries_json or "{}"),
    )


@app.route("/admin/ncd-report/<int:report_id>/delete", methods=["POST"])
@admin_required
def admin_ncd_report_delete(report_id):
    report = NcdReport.query.get_or_404(report_id)
    db.session.delete(report)
    db.session.commit()
    flash("CBHI Form-2 report deleted.", "info")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/ncd-reports/export.csv")
@admin_required
def admin_ncd_reports_export():
    reports = NcdReport.query.order_by(NcdReport.created_at.desc()).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["id", "health_establishment_name", "month", "year", "owner_username", "owner_email", "overall_total", "deaths_total", "created_at"])
    for report in reports:
        _, totals = ncd_rows(report)
        writer.writerow([
            report.id,
            report.health_establishment_name,
            report.month,
            report.year,
            report.ncd_owner.username,
            report.ncd_owner.email,
            totals["overall_total"],
            totals["general_deaths_total"],
            report.created_at.isoformat(),
        ])
    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    response.headers["Content-Disposition"] = "attachment; filename=cbhi_form2_reports.csv"
    return response


@app.route("/admin/reports/consolidated")
@admin_required
def admin_consolidated_reports():
    module = normalized_module_filter(request.args.get("module"))
    cbhi_reports_list = CbhiReport.query.order_by(CbhiReport.created_at.desc()).all()
    ncd_reports_list = NcdReport.query.order_by(NcdReport.created_at.desc()).all()
    rows = build_consolidated_rows(cbhi_reports_list, ncd_reports_list, include_owner=True)
    filtered_rows = filter_consolidated_rows(rows, module)
    return render_template(
        "consolidated_reports.html",
        user=current_user(),
        rows=filtered_rows,
        show_owner=True,
        title="Admin Consolidated Reports",
        selected_module=module,
        module_options=CONSOLIDATED_MODULE_LABELS,
    )


@app.route("/admin/reports/consolidated/print")
@admin_required
def admin_consolidated_reports_print():
    module = normalized_module_filter(request.args.get("module"))
    cbhi_reports_list = CbhiReport.query.order_by(CbhiReport.created_at.desc()).all()
    ncd_reports_list = NcdReport.query.order_by(NcdReport.created_at.desc()).all()
    rows = build_consolidated_rows(cbhi_reports_list, ncd_reports_list, include_owner=True)
    filtered_rows = filter_consolidated_rows(rows, module)
    return render_template(
        "consolidated_reports_print.html",
        rows=filtered_rows,
        show_owner=True,
        title="Admin Consolidated Reports",
        selected_module=module,
        selected_module_label=CONSOLIDATED_MODULE_LABELS[module],
    )


@app.route("/admin/reports/consolidated/export/csv")
@admin_required
def admin_consolidated_reports_csv():
    module = normalized_module_filter(request.args.get("module"))
    cbhi_reports_list = CbhiReport.query.order_by(CbhiReport.created_at.desc()).all()
    ncd_reports_list = NcdReport.query.order_by(NcdReport.created_at.desc()).all()
    rows = build_consolidated_rows(cbhi_reports_list, ncd_reports_list, include_owner=True)
    filtered_rows = filter_consolidated_rows(rows, module)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["report_type", "institution", "district", "month_year", "metric_one_label", "metric_one", "metric_two_label", "metric_two", "owner_username", "owner_email", "created_at"])
    for row in filtered_rows:
        writer.writerow([
            row["report_type"], row["institution"], row["district"], row["month_year"],
            row["metric_one_label"], row["metric_one"], row["metric_two_label"], row["metric_two"],
            row["owner_username"], row["owner_email"], row["created_at"].isoformat(),
        ])

    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    filename = "admin_consolidated_reports.csv" if module == "all" else f"admin_consolidated_reports_{module}.csv"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@app.route("/admin/reports/consolidated/export/excel")
@admin_required
def admin_consolidated_reports_excel():
    module = normalized_module_filter(request.args.get("module"))
    cbhi_reports_list = CbhiReport.query.order_by(CbhiReport.created_at.desc()).all()
    ncd_reports_list = NcdReport.query.order_by(NcdReport.created_at.desc()).all()
    rows = build_consolidated_rows(cbhi_reports_list, ncd_reports_list, include_owner=True)
    filtered_rows = filter_consolidated_rows(rows, module)
    output = build_consolidated_excel(f"Admin Consolidated Reports - {CONSOLIDATED_MODULE_LABELS[module]}", filtered_rows, include_owner=True)

    response = make_response(output.read())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    filename = "admin_consolidated_reports.xlsx" if module == "all" else f"admin_consolidated_reports_{module}.xlsx"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_id", None)
    flash("Admin logged out.", "info")
    return redirect(url_for("admin_login"))


_db_initialized = False


def initialize_database() -> bool:
    global _db_initialized
    with app.app_context():
        _is_sqlite = app.config["SQLALCHEMY_DATABASE_URI"].startswith("sqlite")

        try:
            if _is_sqlite:
                # Migrate hospital_report table if schema is outdated (SQLite only)
                try:
                    result = db.session.execute(text("PRAGMA table_info(hospital_report)")).fetchall()
                    if result:
                        existing_cols = {row[1] for row in result}
                        if "op_new_male" not in existing_cols:
                            db.session.execute(text("DROP TABLE hospital_report"))
                            db.session.commit()
                except Exception:
                    db.session.rollback()

            db.create_all()

            if _is_sqlite:
                # Lightweight migration for existing SQLite DBs created before is_active existed.
                try:
                    result = db.session.execute(text("PRAGMA table_info(user)")).fetchall()
                    user_columns = [row[1] for row in result]
                    if "is_active" not in user_columns:
                        db.session.execute(text("ALTER TABLE user ADD COLUMN is_active BOOLEAN DEFAULT 1"))
                        db.session.execute(text("UPDATE user SET is_active = 1 WHERE is_active IS NULL"))
                        db.session.commit()
                except Exception:
                    db.session.rollback()

            admin_username = os.getenv("ADMIN_USERNAME", "admin")
            admin_email = os.getenv("ADMIN_EMAIL", "admin@nova.local")
            admin_password = os.getenv("ADMIN_PASSWORD", "Admin@123")

            admin = User.query.filter_by(username=admin_username).first()
            if not admin:
                admin = User(username=admin_username, email=admin_email, is_admin=True)
                admin.set_password(admin_password)
                db.session.add(admin)
                db.session.commit()
            else:
                # Keep admin credentials aligned with environment configuration on deploy.
                admin.email = admin_email
                admin.is_admin = True
                admin.set_password(admin_password)
                db.session.commit()

            _db_initialized = True
            return True
        except OperationalError as exc:
            db.session.rollback()
            print(f"[DB INIT WARNING] Database unavailable during startup: {exc}")
            return False
        except Exception as exc:
            db.session.rollback()
            print(f"[DB INIT WARNING] Database initialization failed: {exc}")
            return False


initialize_database()


@app.before_request
def ensure_database_initialized() -> None:
    if not _db_initialized:
        initialize_database()


if __name__ == "__main__":
    app.run(debug=True)
